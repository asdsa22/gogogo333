#!/usr/bin/env python3
"""pet_stats.xlsx → pets_data.json 변환 스크립트
사용법: python convert_excel.py [엑셀경로] [JSON출력경로]
기본값: pet_stats.xlsx → pets_data.json
"""
import sys, json

def convert(xlsx_path="pet_stats.xlsx", json_path="pets_data.json"):
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    pets = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, 1).value
        if not name: continue
        l_val = ws.cell(row, 12).value
        skill = ws.cell(row, 11).value
        pet = {
            "name": str(name).replace(".gif", ""),
            "gif": str(name),
            "hp": round(float(ws.cell(row, 2).value or 0), 1),
            "atk": round(float(ws.cell(row, 3).value or 0), 1),
            "def": round(float(ws.cell(row, 4).value or 0), 1),
            "agi": round(float(ws.cell(row, 5).value or 0), 1),
            "hp_gain": round(float(ws.cell(row, 6).value or 0), 3),
            "atk_gain": round(float(ws.cell(row, 7).value or 0), 3),
            "def_gain": round(float(ws.cell(row, 8).value or 0), 3),
            "agi_gain": round(float(ws.cell(row, 9).value or 0), 3),
            "atk_mul": round(float(ws.cell(row, 10).value or 1), 2),
            "skill": str(skill).strip().upper() if skill and str(skill).strip() else None,
            "total_gain": round(float(l_val), 3) if l_val else 0,
        }
        pets.append(pet)
    wb.close()

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(pets, f, ensure_ascii=False)

    sk = sum(1 for p in pets if p["skill"])
    print(f"✅ 변환 완료: {len(pets)}종 (특수기술 {sk}종) → {json_path}")

if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "pet_stats.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "pets_data.json"
    convert(xlsx, out)
